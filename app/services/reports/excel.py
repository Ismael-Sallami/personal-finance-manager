"""Excel export. One place for the styling and the column autofit."""
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings

_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="F2F6FA")


def _money_format() -> str:
    return f'#,##0.00 "{settings.currency_symbol}"'


def _write_sheet(ws: Worksheet, headers: list[str], rows: list[list]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            if r % 2 == 0:
                cell.fill = _ALT_FILL
            if isinstance(val, (int, float, Decimal)) and c > 1:
                cell.number_format = _money_format()
                cell.alignment = Alignment(horizontal="right")
    # Autofit: widest value in the column plus padding.
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width + 4


def monthly_xlsx(month_name: str, year: int, expenses: list[tuple],
                 income: list[tuple]) -> bytes:
    """expenses/income = list of (category, amount)."""
    wb = Workbook()
    ws_e = wb.active
    ws_e.title = "Expenses"
    _write_sheet(ws_e, ["Category", "Amount"], [[c, float(v)] for c, v in expenses])
    _write_sheet(wb.create_sheet("Income"), ["Category", "Amount"],
                 [[c, float(v)] for c, v in income])
    te = sum(float(v) for _, v in expenses)
    ti = sum(float(v) for _, v in income)
    _write_sheet(wb.create_sheet("Balance"), ["Item", "Amount"],
                 [["Income", ti], ["Expenses", te], ["Balance", ti - te]])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def yearly_xlsx(year: int, by_month: list[list], by_category: list[tuple]) -> bytes:
    """by_month = rows [Month, Income, Expenses, Balance]; by_category = (cat, total)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly summary"
    _write_sheet(ws, ["Month", "Income", "Expenses", "Balance"], by_month)
    _write_sheet(wb.create_sheet("Expenses by category"), ["Category", "Total"],
                 [[c, float(v)] for c, v in by_category])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def net_worth_xlsx(split: list[tuple], kpis: dict) -> bytes:
    """split = (item, value)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Net worth"
    _write_sheet(ws, ["Item", "Value"], [[c, float(v)] for c, v in split])
    _write_sheet(wb.create_sheet("Summary"), ["KPI", "Value"],
                 [[k, float(v)] for k, v in kpis.items()])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
